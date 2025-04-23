import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time


def get_word_info(word):
    """从有道词典抓取单词的音标、中文释义和例句"""
    url = f"http://dict.youdao.com/w/{word}/"
    try:
        response = requests.get(url)
        if response.status_code != 200:
            return None, None, None
        soup = BeautifulSoup(response.text, 'html.parser')

        # 提取音标
        phonetic = soup.find('span', class_='phonetic')
        phonetic_text = phonetic.text if phonetic else ''

        # 提取中文释义
        trans_container = soup.find('div', class_='trans-container')
        definition = ''
        if trans_container:
            ul = trans_container.find('ul')
            if ul:
                li = ul.find('li')
                definition = li.text if li else ''

        # 提取例句
        examples = soup.find('div', class_='examples')
        example_sentence = ''
        if examples:
            p = examples.find('p')
            example_sentence = p.text if p else ''

        return phonetic_text, definition, example_sentence
    except Exception as e:
        print(f"抓取单词 {word} 时出错: {e}")
        return None, None, None


def main():
    """主函数：读取 Excel 文件，添加信息并保存"""
    # 加载 Excel 文件
    wb = load_workbook('words.xlsx')
    sheet = wb['words']

    # 设置表头（可选，如果已有表头可跳过）
    sheet['B1'] = '音标'
    sheet['C1'] = '中文释义'
    sheet['D1'] = '例句'

    # 从第二行开始遍历单词
    for row in sheet.iter_rows(min_row=2, max_col=1):
        word = row[0].value
        if word:
            phonetic, definition, example = get_word_info(word)
            # 写入音标、释义和例句
            sheet.cell(row=row[0].row, column=2).value = phonetic if phonetic else 'N/A'
            sheet.cell(row=row[0].row, column=3).value = definition if definition else 'N/A'
            sheet.cell(row=row[0].row, column=4).value = example if example else 'N/A'
            print(f"已处理: {word}")
            time.sleep(1)  # 暂停 1 秒，避免请求过于频繁

    # 保存更新后的 Excel 文件
    wb.save('words_with_info.xlsx')
    print("处理完成，结果已保存到 words_with_info.xlsx")


if __name__ == "__main__":
    main()