import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time
import shutil
import os
from datetime import datetime

class WordProcessor:
    # 定义列索引常量，便于维护
    WORD_COL = 1  # A列：单词
    PHONETIC_COL = 2  # B列：音标
    DEFINITION_COL = 3  # C列：释义
    EXAMPLE_COL = 4  # D列：例句
    STATUS_COL = 5  # E列：处理状态

    def __init__(self, file_path):
        """初始化 WordProcessor 类，加载 Excel 文件"""
        self.file_path = file_path
        self.wb = load_workbook(file_path)

    def create_backup(self):
        """创建原始文件的备份，文件名带有时间戳"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"{os.path.splitext(self.file_path)[0]}_backup_{timestamp}.xlsx"
        try:
            shutil.copy(self.file_path, backup_path)
            print(f"备份已创建: {backup_path}")
        except Exception as e:
            print(f"创建备份时出错: {e}")
            raise

    def get_word_info(self, word, retries=3):
        """从有道词典获取单词的音标、释义和例句，支持重试机制"""
        url = f"http://dict.youdao.com/w/{word}/"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'
        }
        for attempt in range(retries):
            try:
                response = requests.get(url, headers=headers)
                if response.status_code == 200:
                    soup = BeautifulSoup(response.text, 'html.parser')
                    # 提取音标
                    phonetic = soup.find('span', class_='phonetic')
                    phonetic_text = phonetic.text if phonetic else ''
                    # 提取释义
                    trans_container = soup.find('div', class_='trans-container')
                    definition = ''
                    if trans_container:
                        ul = trans_container.find('ul')
                        if ul:
                            li = ul.find('li')
                            definition = li.text if li else ''
                    # 提取例句
                    examples = soup.find('div', class_='examples')
                    example_sentence = ''
                    if examples:
                        p = examples.find('p')
                        example_sentence = p.text if p else ''
                    return phonetic_text, definition, example_sentence
                else:
                    print(f"请求失败，状态码: {response.status_code}")
            except Exception as e:
                print(f"获取单词 {word} 信息时出错: {e}")
            if attempt < retries - 1:
                print(f"等待重试... ({attempt+1}/{retries})")
                time.sleep(5)
        return None, None, None

    def process_sheet(self, sheet):
        """处理单个工作表，更新单词信息"""
        # 确保状态列表头正确
        if sheet.cell(row=1, column=self.STATUS_COL).value != '处理状态':
            sheet.cell(row=1, column=self.STATUS_COL).value = '处理状态'
        # 从第二行开始处理每一行
        for row in sheet.iter_rows(min_row=2, max_col=self.STATUS_COL):
            word_cell = row[self.WORD_COL - 1]
            status_cell = row[self.STATUS_COL - 1]
            if status_cell.value == '已处理':
                print(f"跳过已处理的单词: {word_cell.value}")
                continue
            if word_cell.value and status_cell.value is None:
                phonetic, definition, example = self.get_word_info(word_cell.value)
                row[self.PHONETIC_COL - 1].value = phonetic if phonetic else 'N/A'
                row[self.DEFINITION_COL - 1].value = definition if definition else 'N/A'
                row[self.EXAMPLE_COL - 1].value = example if example else 'N/A'
                status_cell.value = '已处理'
                print(f"已处理: {word_cell.value}")
                time.sleep(1)  # 避免请求过于频繁

    def process_all_sheets(self):
        """处理工作簿中的所有工作表"""
        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            print(f"开始处理工作表: {sheet_name}")
            self.process_sheet(sheet)

    def save(self):
        """保存工作簿到原始文件"""
        self.wb.save(self.file_path)
        print(f"处理完成，结果已保存到 {self.file_path}")

    def run(self):
        """执行整个处理流程"""
        self.create_backup()
        self.process_all_sheets()
        self.save()

def main():
    file_path = 'words.xlsx'  # 请替换为你的Excel文件路径
    processor = WordProcessor(file_path)
    processor.run()

if __name__ == "__main__":
    main()