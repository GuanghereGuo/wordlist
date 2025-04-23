import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time
import shutil
import os
from datetime import datetime

def get_word_info(word):
    """从有道词典获取单词的音标、释义和例句"""
    url = f"http://dict.youdao.com/w/{word}/"
    try:
        response = requests.get(url)
        if response.status_code != 200:
            return None, None, None
        soup = BeautifulSoup(response.text, 'html.parser')

        # 提取音标
        phonetic = soup.find('span', class_='phonetic')
        phonetic_text = phonetic.text if phonetic else ''

        # 提取释义
        trans_container = soup.find('div', class_='trans-container')
        definition = ''
        if trans_container:
            ul = trans_container.find('ul')
            if ul:
                li = ul.find('li')
                definition = li.text if li else ''

        # 提取例句
        examples = soup.find('div', class_='examples')
        example_sentence = ''
        if examples:
            p = examples.find('p')
            example_sentence = p.text if p else ''

        return phonetic_text, definition, example_sentence
    except Exception as e:
        print(f"获取单词 {word} 信息时出错: {e}")
        return None, None, None

def backup_file(file_path):
    """创建原始文件的备份，文件名带有时间戳"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = f"{os.path.splitext(file_path)[0]}_backup_{timestamp}.xlsx"
    try:
        shutil.copy(file_path, backup_path)
        print(f"备份已创建: {backup_path}")
    except Exception as e:
        print(f"创建备份时出错: {e}")
        raise

def main():
    """处理单词，覆盖原始文件，并在处理前创建备份"""
    file_path = 'words.xlsx'  # 请替换为你的Excel文件路径

    # 创建原始文件的备份
    backup_file(file_path)

    # 加载Excel文件
    wb = load_workbook(file_path)
    sheet = wb['words']  # 假设工作表名为'words'

    # 确保表头正确
    if sheet['E1'].value != '处理状态':
        sheet['E1'] = '处理状态'

    # 从第二行开始处理每一行
    for row in sheet.iter_rows(min_row=2, max_col=5):
        word = row[0].value  # 第A列：单词
        status = row[4].value  # 第E列：处理状态
        if status == '已处理':
            print(f"跳过已处理的单词: {word}")
            continue
        if word and status is None:
            phonetic, definition, example = get_word_info(word)
            # 写入音标、释义和例句
            sheet.cell(row=row[0].row, column=2).value = phonetic if phonetic else 'N/A'
            sheet.cell(row=row[0].row, column=3).value = definition if definition else 'N/A'
            sheet.cell(row=row[0].row, column=4).value = example if example else 'N/A'
            # 更新处理状态
            sheet.cell(row=row[0].row, column=5).value = '已处理'
            print(f"已处理: {word}")
            time.sleep(1)  # 暂停1秒，避免请求过于频繁

    # 覆盖原始文件
    wb.save(file_path)
    print(f"处理完成，结果已保存到 {file_path}")

if __name__ == "__main__":
    main()