import shutil
import os
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import logging


def setup_logging(log_level='INFO'):
    """初始化日志配置"""
    logging.basicConfig(
        filename='word_processor.log',
        level=getattr(logging, log_level.upper()),
        format='%(asctime)s - %(levelname)s - %(message)s'
    )


def create_backup(file_path):
    """创建文件备份"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = os.path.basename(file_path)
    dir_name = os.path.dirname(file_path)
    if not os.path.exists(f"{dir_name}/backup/"):
        os.makedirs(f"{dir_name}/backup/")
    backup_path = (f"{dir_name}/backup/"
                   + f"{file_name}_backup_{timestamp}.xlsx")
    shutil.copy(file_path, backup_path)
    logging.info(f"备份已创建: {backup_path}")
    return backup_path
#
# def set_format(file_path, config):
#     wb = load_workbook(file_path)
#
#     # 设置全局字体
#     # if config['format']['sort']:
#     #     sort_excel(file_path, config)
#
#     font = Font(name=config['format']['font']['name'], size=config['format']['font']['size'])
#     for sheet in wb.worksheets:
#         for row in sheet.iter_rows():
#             for cell in row:
#                 cell.font = font
#                 cell.alignment = Alignment(wrap_text=True)
#
#     # 保存修改
#     wb.save(file_path)
#     logging.info(f"字体已设置: {file_path}")

def set_format(file_path, config): # preserving_colors
    try:
        wb = load_workbook(file_path)
        new_font_name = config['format']['font']['name']
        new_font_size = config['format']['font']['size']
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    # 1. 获取单元格当前的字体对象
                    current_font = cell.font
                    # 2. 创建一个新的Font对象，基于current_font，并只修改name和size
                    # openpyxl的Font对象是不可变的，所以我们需要创建一个新的
                    # 并复制所有现有属性，然后覆盖我们想改变的
                    updated_font = Font(
                        name=new_font_name,
                        size=new_font_size,
                        # 以下是保留原有属性的关键：
                        bold=current_font.bold,
                        italic=current_font.italic,
                        color=current_font.color,  # 保留原有颜色！
                        strike=current_font.strike,
                        underline=current_font.underline,
                        vertAlign=current_font.vertAlign,
                        charset=current_font.charset,
                        scheme=current_font.scheme,
                        family=current_font.family,
                        outline=current_font.outline,
                        shadow=current_font.shadow,
                    )
                    cell.font = updated_font

                    # 设置文本自动换行，这部分没有颜色影响，可以保留
                    cell.alignment = Alignment(wrap_text=True)
        # 保存修改
        wb.save(file_path)
        logging.info(f"字体已设置 (保留原有颜色): {file_path}")
    except Exception as e:
        logging.error(f"处理文件 {file_path} 时发生错误: {e}")