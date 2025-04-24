import sys

from openpyxl import load_workbook
import time
from fetcher import get_word_info
import logging

class WordProcessor:
    def __init__(self, file_path, config):
        self.file_path = file_path
        self.config = config
        self.wb = load_workbook(file_path)
        self.cols = config['columns']

    def process_sheet(self, sheet):
        """处理单个工作表"""
        if sheet.cell(row=1, column=self.cols['status']).value != '处理状态':
            sheet.cell(row=1, column=self.cols['status']).value = '处理状态'
        if sheet.cell(row=1, column=self.cols['wrong_history']).value != '错误历史':
            sheet.cell(row=1, column=self.cols['wrong_history']).value = '错误历史'

        for row_idx in range(2, sheet.max_row + 1):  # 遍历每一行，从第2行开始
            # 确保行数据足够长，如果列不足，插入新列
            if sheet.max_column < self.cols['wrong_history']:
                sheet.insert_cols(self.cols['wrong_history'])

            # 直接通过 sheet.cell 操作单元格，避免 row 的索引问题
            word_cell = sheet.cell(row=row_idx, column=self.cols['word'])
            status_cell = sheet.cell(row=row_idx, column=self.cols['status'])
            wrong_history_cell = sheet.cell(row=row_idx, column=self.cols['wrong_history'])

            if status_cell.value == '已处理':
                logging.info(f"跳过已处理单词: {word_cell.value}")
                continue
            if status_cell.value == '失败':
                logging.info(f"跳过失败单词: {word_cell.value}")
                continue
            if word_cell.value and status_cell.value is None:
                phonetic, definition, example = get_word_info(word_cell.value, self.config)
                sheet.cell(row=row_idx, column=self.cols['phonetic']).value = phonetic if phonetic else 'N/A'
                sheet.cell(row=row_idx, column=self.cols['definition']).value = definition if definition else 'N/A'
                sheet.cell(row=row_idx, column=self.cols['example']).value = example if example else 'N/A'
                if not phonetic or not definition or not example:
                    wrong_history_cell.value = word_cell.value
                    if not phonetic and not definition and not example:
                        status_cell.value = '失败'
                        logging.error(f"处理失败: {word_cell.value}")
                        print(f"处理失败: {word_cell.value}", file=sys.stderr)
                        with open('../failed_words.txt', 'a', encoding='utf-8') as f:
                            f.write(f"{word_cell.value} in {sheet.title}\n")
                        continue
                    else:
                        logging.warning(f"部分成功: {word_cell.value}")
                        print(f"警告: {word_cell.value}", file=sys.stderr)
                status_cell.value = '已处理'
                logging.info(f"已处理: {word_cell.value}")
                print(f"已处理: {word_cell.value}")
                time.sleep(self.config['request']['delay'])

    def process_all_sheets(self):
        """处理所有工作表"""
        for sheet_name in self.wb.sheetnames:
            logging.info(f"开始处理工作表: {sheet_name}")
            print(f"开始处理工作表: {sheet_name}")
            self.process_sheet(self.wb[sheet_name])

    def save(self):
        """保存工作簿"""
        self.wb.save(self.file_path)
        logging.info(f"处理完成，保存到 {self.file_path}")
        print(f"处理完成，保存到 {self.file_path}")